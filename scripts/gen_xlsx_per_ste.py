#!/usr/bin/env python3
"""gen_xlsx_per_ste.py — 1 fichier Excel par société dans DATA/<TICKER>/KPIs.xlsx"""
import json
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = Path("/Users/yann/Desktop/Projets 2025 26/App KPI/DATA")
PIPELINE_DIR = Path("/Users/yann/spx-app/src/data/v2-pipeline")

def sanitize(s):
    """Remove XML-illegal characters from a string."""
    if not isinstance(s, str):
        return s
    # Remove control chars except tab, newline, carriage return
    s = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', s)
    return s.replace("\n", " ").replace("\r", " ")

def make_xlsx(ticker, company_data, dest_folder):
    wb = openpyxl.Workbook()

    # ── Sheet 1: KPIs ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "KPIs"

    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD")
    )

    name = company_data.get("name", ticker)
    sector = company_data.get("sector", "")
    ws["A1"] = f"{name} ({ticker})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = sector
    ws["A2"].font = Font(italic=True, color="666666")

    kpis = company_data.get("kpis", [])

    # Find max history length
    max_hist = max((len(k.get("history") or []) for k in kpis), default=0)

    # Header row at row 4
    headers = ["KPI (FR)", "KPI (EN)", "Unité", "Périodicité", "Dernière date", "YoY"]
    headers += [f"T-{max_hist - 1 - i}" for i in range(max_hist - 1)] + ["Dernière valeur"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for row_idx, kpi in enumerate(kpis, 5):
        fill = alt_fill if row_idx % 2 == 0 else None
        history = kpi.get("history") or []
        # Pad history on the left if shorter than max_hist
        def extract_val(v):
            if isinstance(v, dict):
                return v.get("value")
            return v
        padded = [None] * (max_hist - len(history)) + [extract_val(v) for v in history]

        row_data = [
            sanitize(kpi.get("name_fr", "") or ""),
            sanitize(kpi.get("name_en", "") or ""),
            sanitize(kpi.get("unit", "") or ""),
            kpi.get("period_type", ""),
            kpi.get("last_data_date", ""),
            sanitize(kpi.get("yoy", "") or ""),
        ] + padded

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fill:
                cell.fill = fill
            cell.border = border
            if col > 6 and val is not None:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 8
    for i in range(7, 7 + max_hist):
        ws.column_dimensions[get_column_letter(i)].width = 12

    ws.freeze_panes = "A5"

    # ── Sheet 2: Gouvernance ───────────────────────────────────────────────────
    gov = company_data.get("governance", {})
    if gov:
        wg = wb.create_sheet("Gouvernance")
        wg["A1"] = f"Gouvernance — {name}"
        wg["A1"].font = Font(bold=True, size=13)

        row = 3
        for key, val in gov.items():
            if key.startswith("_"):
                continue
            wg.cell(row=row, column=1, value=key).font = Font(bold=True)
            if isinstance(val, (dict, list)):
                wg.cell(row=row, column=2, value=sanitize(json.dumps(val, ensure_ascii=False)))
            else:
                wg.cell(row=row, column=2, value=sanitize(str(val)) if val is not None else val)
            row += 1

        wg.column_dimensions["A"].width = 30
        wg.column_dimensions["B"].width = 60

    # ── Sheet 3: Infos générales ───────────────────────────────────────────────
    wi = wb.create_sheet("Infos")
    wi["A1"] = "Infos générales"
    wi["A1"].font = Font(bold=True, size=13)
    fields = ["ticker", "name", "sector", "subsector", "tagline", "founded", "ipo"]
    for i, field in enumerate(fields, 3):
        wi.cell(row=i, column=1, value=field).font = Font(bold=True)
        wi.cell(row=i, column=2, value=company_data.get(field, ""))
    wi.column_dimensions["A"].width = 18
    wi.column_dimensions["B"].width = 50

    out_path = dest_folder / "KPIs.xlsx"
    wb.save(out_path)
    return out_path


def main():
    pipeline_files = [f for f in PIPELINE_DIR.glob("*.json")
                      if "gemini" not in f.name and "bak" not in f.name]

    total = len(pipeline_files)
    done = 0
    skipped = 0
    errors = 0

    print(f"Traitement de {total} fichiers...")

    for pf in sorted(pipeline_files):
        try:
            data = json.loads(pf.read_text())
        except Exception as e:
            errors += 1
            continue

        if not isinstance(data, dict):
            skipped += 1
            continue

        ticker = data.get("ticker") or pf.stem.upper()
        kpis = data.get("kpis", [])
        if not kpis:
            skipped += 1
            continue

        # Find the DATA folder — try exact match, then uppercase
        folder = DATA_DIR / pf.stem
        if not folder.exists():
            folder = DATA_DIR / pf.stem.upper()
        if not folder.exists():
            # Create it
            folder = DATA_DIR / pf.stem
            folder.mkdir(parents=True, exist_ok=True)

        out = folder / "KPIs.xlsx"
        if out.exists():
            skipped += 1
            continue

        try:
            make_xlsx(ticker, data, folder)
            done += 1
            if done % 50 == 0:
                print(f"  {done}/{total} créés...")
        except Exception as e:
            print(f"  ERREUR {ticker}: {e}")
            errors += 1

    print(f"\nTerminé: {done} créés, {skipped} ignorés, {errors} erreurs")


if __name__ == "__main__":
    main()
